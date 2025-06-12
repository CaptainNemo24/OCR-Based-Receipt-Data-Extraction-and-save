import os
import requests
import uuid
import time
import json
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openai import OpenAI

# CLOVA OCR API 호출
api_url = 'YOUR_API_URL'
secret_key = 'YOUR_SECRET_KEY'
# 이스케이프 문자 무시 + 포맷 가능성을 위해 fr를 통해 파일 경로 읽기
image_file = fr'YOUR_FILE_FATH\jpg'

request_json = {
    'images': [
        {
            'format': 'jpg',
            'name': 'demo'
        }
    ],
    'requestId': str(uuid.uuid4()),
    'version': 'V2',
    'timestamp': int(round(time.time() * 1000))
}

payload = {'message': json.dumps(request_json).encode('UTF-8')}

# 대량 image file 처리와 원하는 경로에서 파일 로드 및 생성
for filename in os.listdir(image_file):
    if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp')):
        image_path = os.path.join(image_file, filename)
        print(f"[{filename}] 처리 중...")

        with open(image_path, 'rb') as f:
            files = [('file', f)]
            headers = {
            'X-OCR-SECRET': secret_key
            }
            response = requests.request("POST", api_url, headers=headers, data = payload, files = files)
            json_filename = os.path.splitext(filename)[0] + ".json"
            json_data = os.path.join(r"YOUR_FILE_FATH\json", json_filename)
            print("최종 파일 경로:", json_data)
        # json 파일이 있을 경우에는 불러오고, 없을 경우에는 새로 생성
        try:
            with open(json_data, 'r', encoding='utf-8') as f:
                data = json.load(f)
                fields = data['images'][0]['fields']
                print(f"{json_data} 로드 완료")
        except FileNotFoundError:
            with open(json_data, 'w', encoding='utf-8') as f:
                json.dump(response.json(), f, ensure_ascii=False, indent=4)
                fields = response.json()['images'][0]['fields']
                print(f"{json_data} 생성 완료")
        
    # 응답 내용을 영수증 형태로 다시 변환
    string_result = ''
    for i in fields:
        if i['lineBreak'] == True:
            linebreak = '\n'
        else:
            linebreak = ' '
        string_result = string_result + i['inferText'] + linebreak
        
    print(string_result)

    # OpenAI prompt 응답 생성
    client = OpenAI(
    api_key="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"
    )

    completion = client.chat.completions.create(
    model="gpt-3.5-turbo-1106",
    messages=[
        {"role": "system", "content": "너는 영수증에서 날짜, 업체명, 품목, 단가, 수량, 금액을 찾아 분석하고 JSON 파일로 만들 수 있어. 모든 JSON 입력에서 상품 리스트 역할을 하는 키를 찾아서, 그걸 '상품목록'이라는 키로 변환한 JSON을 출력하지."},
        {"role": "user", "content": f"{string_result}를 분석해서 날짜(시간 제외), 업체명, 품목, 단가, 수량, 금액만 추출해서 정리해줘. 모두 문자열이지. 만약 상품명이 숫자로만 되어있으면 그건 제외해줘. 그리고 매장명 -> 업체명, 매출일 -> 날짜로 변경해줘. 단가, 금액에 .은 ,로 인식해줘."}
    ]
    )
    prompt = completion.choices[0].message.content

    print(prompt)

    data = json.loads(prompt)

    # 날짜와 업체명 추출
    sales_date = data["날짜"]
    store_name = data["업체명"]

    # 파일 경로 및 오픈할 sheet 이름
    file_path = r"YOUR_FILE_FATH\csv\샘플 데이터.xlsx"
    sheet_name = "지출내역"

    #데이터 프레임으로 전환 및 생성
    receipt_data = pd.DataFrame(data['상품목록'])

    # 날짜, 업체명을 컬럼으로 앞에 삽입
    receipt_data.insert(0, "날짜", sales_date)
    receipt_data.insert(1, "업체명", store_name)

    # 간혹 날짜 형식이 yy-mm-dd 되어 있는 경우가 있으므로 날짜 형식 변환: yy-mm-dd → yyyy-mm-dd
    def convert_yy_to_yyyy(date_str):
        # 정규표현식: 정확히 yy-mm-dd 형식만 매칭
        if re.fullmatch(r"\d{2}-\d{2}-\d{2}", date_str):
            return pd.to_datetime(date_str, format="%y-%m-%d").strftime("%Y-%m-%d")
        # 변환하지 않음
        else:
            return pd.to_datetime(date_str)
    
    receipt_data["날짜"] = receipt_data["날짜"].apply(convert_yy_to_yyyy)

    # 쉼표나 공백 등 제거하고 숫자로 변환
    columns_to_convert = ["단가", "수량", "금액"]
    
    for col in columns_to_convert:
        receipt_data[col] = receipt_data[col].astype(str).str.replace(",", "").str.strip().fillna("0").astype(int)

    # 엑셀 열기
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # 테이블 정보 가져오기
    table_name = list(ws.tables.keys())[0]
    table = ws.tables[table_name]
    start_cell, end_cell = table.ref.split(":")
    start_col_letter = ''.join(filter(str.isalpha, start_cell))
    start_row = int(''.join(filter(str.isdigit, start_cell)))
    end_col_letter = ''.join(filter(str.isalpha, end_cell))

    start_col_index = column_index_from_string(start_col_letter)
    end_col_index = column_index_from_string(end_col_letter)

    # 실제 데이터가 있는 마지막 행 찾기
    def get_last_data_row(ws, start_row, col_index):
        row = start_row
        while ws.cell(row=row, column=col_index).value:
            row += 1
        return row - 1

    actual_last_row = get_last_data_row(ws, start_row + 1, start_col_index)

    # 새 데이터 삽입 (실제 마지막 데이터 아래부터)
    for r_idx, row in enumerate(receipt_data.values.tolist(), start=actual_last_row + 1):
        for c_idx, value in enumerate(row):
            ws.cell(row=r_idx, column=start_col_index + c_idx, value=value)

    # 테이블 범위 재설정
    new_end_row = actual_last_row + len(receipt_data)
    new_ref = f"{start_col_letter}{start_row}:{end_col_letter}{new_end_row}"

    # 지정한 파일 경로에 저장 및 확인
    wb.save(file_path)
    print(f"{file_path} 파일에 저장되었습니다.")
